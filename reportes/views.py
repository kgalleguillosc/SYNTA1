from django.http import HttpResponse
from django.views.generic.base import TemplateView
from django.views.generic import ListView
from django.core.files.base import ContentFile
from django.utils import timezone
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from trabajadores.models import Perfil
from turnos.models import Turno
from reportes.models import  ReporteGenerado 

# Create your views here.
class ReportePersonasExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        # Obtenemos todos los perfiles desde la base de datos
        perfiles = Perfil.objects.all()

        # Creamos un nuevo libro de trabajo (Excel)
        wb = Workbook()
        ws = wb.active

        # Título del reporte
        ws['B1'] = 'USUARIOS SYNTA'
        ws.merge_cells('B1:G1')  # Aumentamos las celdas fusionadas para más columnas

        # Creamos los encabezados desde la fila 3
        ws['A3'] = 'RUT'
        ws['B3'] = 'NOMBRE'
        ws['C3'] = 'APELLIDO PATERNO'
        ws['D3'] = 'APELLIDO MATERNO'
        ws['E3'] = 'NOMBRE DE USUARIO'
        ws['F3'] = 'EMAIL'
        ws['G3'] = 'ROL'
        ws['H3'] = 'DEPARTAMENTO'
        ws['I3'] = 'SUPERVISOR'
        ws['J3'] = 'CONTRASEÑA'

        fila = 4

        # Recorremos el conjunto de perfiles y vamos escribiendo cada dato en las celdas
        for perfil in perfiles:
            ws.cell(row=fila, column=1).value = perfil.rut  # RUT
            ws.cell(row=fila, column=2).value = perfil.first_name  # Nombre
            ws.cell(row=fila, column=3).value = perfil.last_name  # Apellido paterno
            ws.cell(row=fila, column=4).value = perfil.last_name  # Apellido materno (Si tienes otro campo, reemplazar)
            ws.cell(row=fila, column=5).value = perfil.username  # Nombre de usuario
            ws.cell(row=fila, column=6).value = perfil.email  # Email
            ws.cell(row=fila, column=7).value = perfil.get_rol_display()  # Rol (desplegado como texto legible)
            ws.cell(row=fila, column=8).value = perfil.departamento  # Departamento
            ws.cell(row=fila, column=9).value = f"{perfil.supervisor.first_name} {perfil.supervisor.last_name}" if perfil.supervisor else "N/A"
            ws.cell(row=fila, column=10).value = 'N3r!oT1x#L9dZp'

            fila += 1

        # Guardamos el archivo Excel en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Establecemos el nombre del archivo
        nombre_archivo = "UsuariosSyntaExcel.xlsx"

        # Definimos que el tipo de respuesta a devolver es un archivo de Microsoft Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"

        # Guardar el archivo en el modelo ReporteGenerado
        reporte_obj = ReporteGenerado(
            usuario=request.user,  # Guardamos el usuario que generó el reporte
            tipo_reporte="usuarios",
            departamento="",  
            fecha_inicio=None,  
            fecha_fin=None,  
            fecha_generacion=timezone.now(),
        )
        reporte_obj.archivo_reporte.save(nombre_archivo, ContentFile(excel_file.getvalue()))
        reporte_obj.save()

        # Enviar el archivo al usuario
        response.write(excel_file.getvalue())

        return response


# Diccionario para mapear el código de turno al nombre de turno
TURNO_TIPOLOGIA = {
    'TE51.2': '41',
    'TE51': '40',
    'TE39': '39',
    'A41': '38',
    'A60': '37',
    'P28': '36',
    'P22': '35',
    'A2.B': '34',
    'A18.A': '28',
    'A36.A.2': '33',
    'A36(TE51)': '32',
    'F54.1': '31',
    'P07(A60)': '30',
    'F52(A41)': '29',
    'A13.A': '27',
    'A2.A': '26',
    'A8': '22',
    'B14': '16',
    'A36': '7',
    'A2': '1',
    'A3': '2',
    'A10': '3',
    'A13': '4',
    'A14': '5',
    'A57': '8',
    'F54': '9',
    'P37': '11',
    'P50': '12',
    'P07': '10',
    'A18': '6',
    'B10': '17',
    'A7': '21',
    'P07.1': '13',
    'A36.1': '14',
    'PRACTICANTE': '20',
    'F52': '23'
}


class ReporteTurnosExcelTALANA(TemplateView):

    def get(self, request, *args, **kwargs):
        departamento = request.GET.get('departamento')
        trabajador_rut = request.GET.get('trabajador')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # Obtener los turnos
        turnos = Turno.objects.filter(estado="APROBADO").order_by('fecha_inicio')
        # Filtrar por departamento si está definido
        if departamento:
            trabajadores_ids = Perfil.objects.filter(departamento=departamento).values_list('rut', flat=True)
            turnos = turnos.filter(rut_trabajador__in=trabajadores_ids)

        # Filtrar por rut del trabajador si está definido
        if trabajador_rut:
            turnos = turnos.filter(rut_trabajador=trabajador_rut)

        if fecha_inicio:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                turnos = turnos.filter(fecha_inicio__gte=fecha_inicio)
            except ValueError:
                return HttpResponse("Formato de fecha_inicio inválido. Debe ser YYYY-MM-DD.", status=400)
        if fecha_fin:
            try:
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                turnos = turnos.filter(fecha_fin__lte=fecha_fin)
            except ValueError:
                return HttpResponse("Formato de fecha_fin inválido. Debe ser YYYY-MM-DD.", status=400)

        # Crear un nuevo libro de trabajo (Excel)
        wb = Workbook()
        ws = wb.active

        # Encabezados de las columnas
        ws['A1'] = 'RUT'
        ws['B1'] = 'TURNO'
        ws['C1'] = 'DESDE'
        ws['D1'] = 'HASTA'
        ws['E1'] = 'DÍA DE INICIO'

        # Inicializamos la fila de los datos
        fila = 2

        # Recorremos el conjunto de turnos y escribimos cada dato en las celdas
        for turno in turnos:
            # Obtenemos el perfil asociado al turno
            perfil = Perfil.objects.get(rut=turno.rut_trabajador)
            rut_sin_puntos_ni_guiones = perfil.rut.replace(".", "")

            # Obtener el nombre del turno basado en su código
            nombre_turno = TURNO_TIPOLOGIA.get(turno.codigo_turno, 'Turno Desconocido')

            # Rellenamos los datos del reporte
            ws.cell(row=fila, column=1).value = rut_sin_puntos_ni_guiones
            ws.cell(row=fila, column=2).value = nombre_turno  # Nombre del turno (mapeado desde el diccionario)
            ws.cell(row=fila, column=3).value = turno.fecha_inicio.strftime('%d-%m-%Y')  # Desde (fecha de inicio)
            ws.cell(row=fila, column=4).value = turno.fecha_fin.strftime('%d-%m-%Y')  # Hasta (fecha de fin)
            ws.cell(row=fila, column=5).value = 1  # DÍA DE INICIO, ajusta esto según tu lógica

            fila += 1

        # Establecemos el nombre del archivo
        nombre_archivo = "ReporteTurnosTALANA.xlsx"

        # Guardamos el archivo Excel en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Definimos que el tipo de respuesta a devolver es un archivo de Microsoft Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"

        # Guardar el archivo en el modelo ReporteGenerado
        reporte_obj = ReporteGenerado(
            usuario=request.user,  # Aquí guardamos el usuario que generó el reporte
            tipo_reporte="talana",
            departamento=departamento,
            fecha_inicio=fecha_inicio if fecha_inicio else None,
            fecha_fin=fecha_fin if fecha_fin else None,
            fecha_generacion=timezone.now(),
        )
        reporte_obj.archivo_reporte.save(nombre_archivo, ContentFile(excel_file.getvalue()))
        reporte_obj.save()

        # Enviar el archivo al usuario
        response.write(excel_file.getvalue())

        return response



class ReporteTurnosExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        departamento = request.GET.get('departamento')
        trabajador_rut = request.GET.get('trabajador')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # Obtener los turnos
        turnos = Turno.objects.filter(estado="APROBADO").order_by('fecha_inicio')


        # Filtrar por departamento si está definido
        if departamento:
            trabajadores_ids = Perfil.objects.filter(departamento=departamento).values_list('rut', flat=True)
            turnos = turnos.filter(rut_trabajador__in=trabajadores_ids)

        # Filtrar por rut del trabajador si está definido
        if trabajador_rut:
            turnos = turnos.filter(rut_trabajador=trabajador_rut)

        if fecha_inicio:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                turnos = turnos.filter(fecha_inicio__gte=fecha_inicio)
            except ValueError:
                return HttpResponse("Formato de fecha_inicio inválido. Debe ser YYYY-MM-DD.", status=400)
        if fecha_fin:
            try:
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                turnos = turnos.filter(fecha_fin__lte=fecha_fin)
            except ValueError:
                return HttpResponse("Formato de fecha_fin inválido. Debe ser YYYY-MM-DD.", status=400)

        # Crear un nuevo archivo Excel
        wb = Workbook()
        ws = wb.active

        # Título del reporte
        ws['A1'] = 'REPORTE DE TURNOS APROBADOS'
        ws.merge_cells('A1:F1')

        # Encabezados de las columnas
        ws['A3'] = 'RUT'
        ws['B3'] = 'NOMBRE COMPLETO'
        ws['C3'] = 'CÓDIGO DE TURNO'
        ws['D3'] = 'FECHA DE INICIO'
        ws['E3'] = 'FECHA DE FIN'
        ws['F3'] = 'DEPARTAMENTO'
        ws['G3'] = 'SUPERVISOR'

        # Inicializamos la fila de los datos
        fila = 4

        # Rellenar los datos del reporte
        for turno in turnos:
            perfil = Perfil.objects.get(rut=turno.rut_trabajador)
            rut_sin_puntos_ni_guiones = perfil.rut.replace(".", "")
    

            ws.cell(row=fila, column=1).value = rut_sin_puntos_ni_guiones
            ws.cell(row=fila, column=2).value = f"{perfil.first_name} {perfil.last_name}"
            ws.cell(row=fila, column=3).value = turno.codigo_turno
            ws.cell(row=fila, column=4).value = turno.fecha_inicio.strftime('%Y-%m-%d')
            ws.cell(row=fila, column=5).value = turno.fecha_fin.strftime('%Y-%m-%d')
            ws.cell(row=fila, column=6).value = perfil.departamento
            ws.cell(row=fila, column=7).value = f"{perfil.supervisor.first_name} {perfil.supervisor.last_name}" if perfil.supervisor else "N/A"

            fila += 1

         # Guardar el archivo Excel en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Crear el nombre del archivo y definir la respuesta HTTP
        nombre_archivo = "ReporteTurnosExcel.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
        response.write(excel_file.getvalue())

        # Guardar el archivo en el modelo ReporteGenerado
        reporte_obj = ReporteGenerado(
            tipo_reporte="turnos",
            usuario= request.user,
            departamento=departamento,
            fecha_inicio=fecha_inicio if fecha_inicio else None,
            fecha_fin=fecha_fin if fecha_fin else None,
            fecha_generacion=timezone.now(),
        )
        reporte_obj.archivo_reporte.save(nombre_archivo, ContentFile(excel_file.getvalue()))
        reporte_obj.save()

        return response
    


# class HistorialReportesView(TemplateView):
#     template_name = 'Reportes.html'  

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['reportes'] = ReporteGenerado.objects.all() 
#         return context